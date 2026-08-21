"""Bounded extraction of source facts with real page, cell, and row provenance."""

from __future__ import annotations

import csv
from pathlib import Path

from atomic_claim_evidence import NUMBER_PATTERN


class EvidenceExtractorReal:
    def __init__(self, max_pdf_pages: int = 100, max_sheets: int = 20, max_rows: int = 2000, max_columns: int = 100, max_facts: int = 500) -> None:
        self.max_pdf_pages = max_pdf_pages
        self.max_sheets = max_sheets
        self.max_rows = max_rows
        self.max_columns = max_columns
        self.max_facts = max_facts

    def extract_pdf_with_pages(self, pdf_path: str) -> list[dict]:
        import pdfplumber

        facts = []
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) > self.max_pdf_pages:
                raise ValueError(f"PDF exceeds the {self.max_pdf_pages}-page processing limit")
            for index, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                for match in NUMBER_PATTERN.finditer(text):
                    start, end = max(0, match.start() - 120), min(len(text), match.end() + 120)
                    facts.append({"claim": match.group(0).strip(), "claim_span": match.group(0).strip(), "source_file": Path(pdf_path).name, "source_type": "pdf", "page_number": index, "source_text": text[start:end].replace("\n", " "), "char_start": match.start(), "char_end": match.end(), "verification_status": "EXTRACTED_WITH_PAGE", "can_use_in_deck": True})
                    if len(facts) >= self.max_facts:
                        return facts
        return facts

    def extract_excel_with_cells(self, excel_path: str) -> list[dict]:
        import openpyxl

        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        formula_book = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
        facts = []
        try:
            for name in workbook.sheetnames[: self.max_sheets]:
                sheet = workbook[name]
                formula_sheet = formula_book[name]
                headers = {}
                for row in sheet.iter_rows(min_row=1, max_row=self.max_rows, max_col=self.max_columns):
                    for cell in row:
                        if cell.value is None:
                            continue
                        text = str(cell.value).strip()
                        if cell.row == 1:
                            headers[cell.column] = text
                        if len(text) > 250 or not list(NUMBER_PATTERN.finditer(text)):
                            continue
                        label = headers.get(cell.column, "")
                        formula_cell = formula_sheet[cell.coordinate]
                        formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
                        context = f"{name}!{cell.coordinate} {label} = {text}".strip()
                        facts.append({"claim": text, "claim_span": text, "source_file": Path(excel_path).name, "source_type": "excel", "sheet": name, "cell": cell.coordinate, "cell_range": f"{name}!{cell.coordinate}", "source_text": context, "formula": formula, "verification_status": "EXTRACTED_WITH_CELL", "can_use_in_deck": True})
                        if len(facts) >= self.max_facts:
                            return facts
        finally:
            workbook.close()
            formula_book.close()
        return facts

    def extract_csv_with_rows(self, csv_path: str) -> list[dict]:
        facts = []
        with open(csv_path, newline="", encoding="utf-8-sig") as source:
            reader = csv.reader(source)
            headers = next(reader, [])
            for row_index, row in enumerate(reader, start=2):
                if row_index > self.max_rows:
                    break
                for column_index, value in enumerate(row[: self.max_columns], start=1):
                    value = value.strip()
                    if len(value) > 250 or not list(NUMBER_PATTERN.finditer(value)):
                        continue
                    heading = headers[column_index - 1] if len(headers) >= column_index else f"column {column_index}"
                    location = f"row {row_index}, column {heading}"
                    facts.append({"claim": value, "source_file": Path(csv_path).name, "source_type": "csv", "cell": f"R{row_index}C{column_index}", "cell_range": location, "source_text": f"{heading} {location} = {value}", "verification_status": "EXTRACTED_WITH_ROW", "can_use_in_deck": True})
                    if len(facts) >= self.max_facts:
                        return facts
        return facts

    def extract_docx_with_paragraph(self, docx_path: str) -> list[dict]:
        import docx

        facts = []
        document = docx.Document(docx_path)
        for index, paragraph in enumerate(document.paragraphs, start=1):
            text = paragraph.text
            for match in NUMBER_PATTERN.finditer(text):
                facts.append({"claim": match.group(0).strip(), "source_file": Path(docx_path).name, "source_type": "docx", "paragraph_number": index, "source_text": text[max(0, match.start() - 120) : match.end() + 120], "verification_status": "EXTRACTED_WITH_PARA", "can_use_in_deck": True})
                if len(facts) >= self.max_facts:
                    return facts
        return facts
